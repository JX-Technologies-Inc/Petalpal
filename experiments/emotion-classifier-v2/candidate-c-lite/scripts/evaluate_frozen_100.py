#!/usr/bin/env python3
"""One-time, read-only evaluation of the locked checkpoint on Frozen 100."""

from __future__ import annotations

import argparse
import json
import re
import subprocess
from pathlib import Path

import numpy as np
import torch
from openpyxl import load_workbook
from sklearn.metrics import precision_recall_fscore_support
from transformers import AutoModelForSequenceClassification, AutoTokenizer


LABELS = [
    "admiration", "amusement", "anger", "annoyance", "approval", "caring",
    "confusion", "curiosity", "disappointment", "disapproval", "disgust",
    "excitement", "fear", "gratitude", "joy", "love", "neutral", "optimism",
    "remorse", "sadness", "surprise",
]
THRESHOLD = 0.5
FROZEN_PATH = Path("/Users/xingranma/Desktop/ml/PetalPal_frozen_100sample.xlsx")
VARIANT_LABELS = [label for label in LABELS if label not in {"neutral", "approval", "disapproval"}]


def normalized_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def find_column(headers: list[object], *names: str) -> int:
    indexed = {normalized_header(value): index for index, value in enumerate(headers)}
    for name in names:
        if normalized_header(name) in indexed:
            return indexed[normalized_header(name)]
    raise ValueError(f"Missing required column; expected one of: {', '.join(names)}")


def parse_labels(value: object) -> list[str]:
    if value is None or str(value).strip() == "":
        return []
    labels = [item.strip().lower() for item in re.split(r"[,;|]", str(value)) if item.strip()]
    unknown = sorted(set(labels) - set(LABELS))
    if unknown:
        raise ValueError(f"Unknown gold labels: {unknown}")
    return labels


def parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    normalized = str(value or "").strip().lower()
    if normalized in {"true", "1", "yes"}:
        return True
    if normalized in {"false", "0", "no"}:
        return False
    raise ValueError(f"Invalid boolean value: {value!r}")


def parse_count(value: object) -> int:
    try:
        result = int(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"Invalid output count: {value!r}") from error
    if result not in {0, 1, 2}:
        raise ValueError(f"Output count must be 0, 1, or 2: {value!r}")
    return result


def read_frozen(path: Path) -> tuple[list[str], list[dict]]:
    if "frozen" not in path.name.lower():
        raise ValueError("Refusing to evaluate a file whose name does not contain 'frozen'")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    journal_col = find_column(headers, "journal", "event/journal", "event or journal", "daily grow journal")
    gold_col = find_column(
        headers,
        "expectedSecondaryEmotions",
        "expected secondary emotions",
        "gold secondary emotions",
        "best secondary emotions",
        "best secondary emotion(s)",
    )
    acceptable_col = find_column(headers, "acceptableAlternatives", "acceptable alternatives", "acceptable alternative(s)")
    wrong_col = find_column(headers, "clearlyWrongEmotions", "clearly wrong emotions", "clearly wrong emotion(s)")
    abstain_col = find_column(headers, "preferNoSecondaryEmotion", "better to return none?")
    redundant_col = find_column(headers, "primaryRedundantEmotions", "emotions redundant with primary", "emotion(s) redundant with primary")
    min_col = find_column(headers, "expectedOutputMin", "min output")
    max_col = find_column(headers, "expectedOutputMax", "max output")
    primary_col = find_column(headers, "primaryGardenMood", "primary mood")
    id_col = find_column(headers, "id")
    texts, annotations = [], []
    for row_number, row in enumerate(rows, start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        journal = str(row[journal_col] or "").strip()
        expected = parse_labels(row[gold_col])
        acceptable = parse_labels(row[acceptable_col])
        wrong = parse_labels(row[wrong_col])
        redundant = parse_labels(row[redundant_col])
        minimum, maximum = parse_count(row[min_col]), parse_count(row[max_col])
        if minimum > maximum:
            raise ValueError(f"Min output exceeds max at worksheet row {row_number}")
        texts.append(journal)
        annotations.append({
            "id": str(row[id_col]),
            "primaryGardenMood": str(row[primary_col]).strip(),
            "expected": expected,
            "acceptable": acceptable,
            "clearlyWrong": wrong,
            "preferNone": parse_bool(row[abstain_col]),
            "redundant": redundant,
            "min": minimum,
            "max": maximum,
        })
    workbook.close()
    if len(texts) != 100:
        raise ValueError(f"Expected exactly 100 Frozen examples, found {len(texts)}")
    return texts, annotations


def label_positions(model) -> list[int]:
    id2label = {int(index): str(label).lower() for index, label in model.config.id2label.items()}
    positions = []
    for wanted in LABELS:
        matches = [index for index, label in id2label.items() if label == wanted]
        if not matches:
            raise ValueError(f"Checkpoint is missing label: {wanted}")
        positions.append(matches[0])
    if model.config.num_labels != 28:
        raise ValueError(f"Expected locked 28-output checkpoint, found {model.config.num_labels}")
    return positions


def predict(texts: list[str], checkpoint: Path, batch_size: int) -> np.ndarray:
    tokenizer = AutoTokenizer.from_pretrained(checkpoint, local_files_only=True)
    model = AutoModelForSequenceClassification.from_pretrained(checkpoint, local_files_only=True).eval()
    positions = label_positions(model)
    device = torch.device("cuda" if torch.cuda.is_available() else "mps" if torch.backends.mps.is_available() else "cpu")
    model.to(device)
    probabilities = np.zeros((len(texts), len(LABELS)), dtype=np.float32)
    nonempty = [(index, text) for index, text in enumerate(texts) if text]
    with torch.inference_mode():
        for start in range(0, len(nonempty), batch_size):
            batch = nonempty[start:start + batch_size]
            encoded = tokenizer(
                [text for _, text in batch], padding=True, truncation=True,
                max_length=128, return_tensors="pt",
            )
            logits = model(**{key: value.to(device) for key, value in encoded.items()}).logits[:, positions]
            scores = torch.sigmoid(logits).cpu().numpy()
            for (index, _), row in zip(batch, scores):
                probabilities[index] = row
    return probabilities


def select_variants(probabilities: np.ndarray, annotations: list[dict], root: Path) -> list[list[str]]:
    payload = [
        {
            "primaryGardenMood": annotation["primaryGardenMood"],
            "candidates": [
                {"label": label, "score": float(row[index])}
                for index, label in enumerate(LABELS)
                if row[index] >= THRESHOLD
            ],
        }
        for row, annotation in zip(probabilities, annotations)
    ]
    selector = (root.parents[1] / "lib/secondary-emotion-selector.js").resolve().as_uri()
    program = f'''import {{ selectFlowerSecondaryEmotions }} from {json.dumps(selector)};
const chunks = []; for await (const chunk of process.stdin) chunks.push(chunk);
const rows = JSON.parse(chunks.join(""));
console.log(JSON.stringify(rows.map(row => selectFlowerSecondaryEmotions(row).map(item => item.label))));'''
    completed = subprocess.run(
        ["node", "--input-type=module", "-e", program], input=json.dumps(payload),
        text=True, capture_output=True, check=True,
    )
    return json.loads(completed.stdout)


def metric_report(truth: np.ndarray, predicted: np.ndarray) -> dict:
    report = {}
    for average in ("macro", "micro"):
        precision, recall, f1, _ = precision_recall_fscore_support(
            truth, predicted, average=average, zero_division=0,
        )
        report[average] = {"precision": float(precision), "recall": float(recall), "f1": float(f1)}
    precision, recall, f1, support = precision_recall_fscore_support(
        truth, predicted, average=None, zero_division=0,
    )
    report["perLabel"] = {
        label: {
            "precision": float(precision[index]),
            "recall": float(recall[index]),
            "f1": float(f1[index]),
            "support": int(support[index]),
        }
        for index, label in enumerate(VARIANT_LABELS)
    }
    return report


def product_report(annotations: list[dict], outputs: list[list[str]]) -> dict:
    output_total = strict_hits = acceptable_hits = wrong_hits = redundant_hits = 0
    required = useful = unwanted_abstentions = 0
    prefer_none = correct_abstentions = exact = acceptable_matches = 0
    counts = {"0": 0, "1": 0, "2": 0}
    for annotation, output in zip(annotations, outputs):
        selected = set(output)
        best = set(annotation["expected"])
        allowed = best | set(annotation["acceptable"])
        output_total += len(output)
        strict_hits += len(selected & best)
        acceptable_hits += len(selected & allowed)
        wrong_hits += len(selected & set(annotation["clearlyWrong"]))
        redundant_hits += len(selected & set(annotation["redundant"]))
        counts[str(len(output))] += 1
        exact += selected == best
        acceptable_matches += annotation["min"] <= len(output) <= annotation["max"] and selected <= allowed
        if annotation["min"] > 0:
            required += 1
            useful += bool(selected & allowed)
            unwanted_abstentions += not output
        if annotation["preferNone"]:
            prefer_none += 1
            correct_abstentions += not output

    ratio = lambda numerator, denominator: numerator / denominator if denominator else 0.0
    return {
        "strictSecondaryPrecision": ratio(strict_hits, output_total),
        "acceptableSecondaryPrecision": ratio(acceptable_hits, output_total),
        "usefulCoverage": ratio(useful, required),
        "correctAbstentionRate": ratio(correct_abstentions, prefer_none),
        "unwantedAbstentionRate": ratio(unwanted_abstentions, required),
        "clearlyWrongEmotionRate": ratio(wrong_hits, output_total),
        "primaryRedundancyRate": ratio(redundant_hits, output_total),
        "outputCountDistribution": counts,
        "exactExpectedMatchRate": ratio(exact, len(outputs)),
        "acceptableMatchRate": ratio(acceptable_matches, len(outputs)),
    }


def main() -> None:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser()
    parser.add_argument("--batch-size", type=int, default=8)
    args = parser.parse_args()

    checkpoint = root / "candidate-c-lite/fine-tuned-v1/best-checkpoint"
    texts, annotations = read_frozen(FROZEN_PATH)
    probabilities = predict(texts, checkpoint, args.batch_size)
    outputs = select_variants(probabilities, annotations, root)
    truth = np.asarray([
        [int(label in annotation["expected"]) for label in VARIANT_LABELS]
        for annotation in annotations
    ], dtype=np.int8)
    predicted = np.asarray([
        [int(label in output) for label in VARIANT_LABELS]
        for output in outputs
    ], dtype=np.int8)
    result = {
        "evaluation": "Frozen 100 one-time final evaluation",
        "checkpoint": str(checkpoint),
        "examples": len(texts),
        "threshold": THRESHOLD,
        "labels": VARIANT_LABELS,
        "classificationMetrics": metric_report(truth, predicted),
        "productMetrics": product_report(annotations, outputs),
    }
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
